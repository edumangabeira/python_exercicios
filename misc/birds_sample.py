# -*-coding:utf-8-*-
from openpyxl import Workbook
# from openpyxl import load_workbook
import random as rd
from string import ascii_uppercase

'''
Generates random values for lables of a pseudo-sample. For this example, I'm using
the number and types of spotted birds while going for a walk in the park as the sample.
The bird names/types will compose the labels(species).
It will also write the labels and frequencies on a spreadsheet(xlsx).
'''


class BirdSamples:

    def __init__(self):
        # self.species = species
        self.species = []
        self.sample = {}

    def read_bird_entries(self):
        # arq_name = input("digite o nome do arquivo com o nome das espécies:")
        arq_name = 'birds.txt'
        bird_list = []
        with open(arq_name, 'r') as arq:
            bird_list = [bird.strip() for bird in arq]
        self.species = bird_list
        print(bird_list)

    def generate_frequency(self):
        for bird in self.species:
            self.sample[bird] = rd.randint(1, 50)
        return self.sample

    def decrp_dict(self, dictionary):
        species = []
        frequency = []

        for key, value in dictionary.items():
            species.append(key)
            frequency.append(value)

        return species, frequency

    def write_excel(self, columns, rows):
        # wb = load_workbook("birds.xlsx")
        # ws = wb.get_sheet_by_name("Sheet1")
        wb = Workbook()
        ws = wb.active
        if len(columns) < 23:
            labels = list(ascii_uppercase)

        for col in range(len(columns)):
            ws['{}{}'.format(labels[col + 1], 1)] = columns[col]
        for row in range(len(rows)):
            ws['{}{}'.format(labels[row + 1], 2)] = rows[row]
        wb.save("birds.xlsx")

    def generate_sample(self):
        self.read_bird_entries()
        sample = self.generate_frequency()
        columns, rows = self.decrp_dict(sample)
        self.write_excel(columns, rows)


if __name__ == '__main__':

    # species = ["cuco", "picapau", "sabiá", "peru", "avestruz", "pombo"]
    # species =
    bird = BirdSamples()
    bird.generate_sample()
